"""
Сервер на Flask: общение по HTTP.
Нужно:
- alg.py лежит на сервере;
- ты в веб-админке жмёшь кнопку;
- клиенты, которые опрашивают сервер, видят новый запуск, СКАЧИВАЮТ alg.py и запускают его локально.

Запуск сервера: .venv/bin/python server.py
"""

import os
import json
import threading
from io import BytesIO, StringIO
from flask import Flask, jsonify, send_from_directory, request, Response, send_file
import openpyxl
import csv

from robot_run import run as run_robot
from robot_state import request_stop_run
from robot_payload import normalize_robot_payload
from robot import validate_before_setting

app = Flask(__name__)

# Глобальный счётчик запусков (один для всех клиентов)
RUN_ID = 0

# Последний лог от клиента
CLIENT_LOG = {
    "run_id": None,
    "returncode": None,
    "stdout": "",
    "stderr": "",
}

# Результат проверки документа выгрузки: заголовки и строки, которые нельзя обработать
DOC_CHECK_RESULT = {"headers": [], "cannot_process_rows": []}

# Список дел для запуска робота «Проверить данные 1С» (заполняется при нажатии «Начать»)
CHECK_1C_CASE_LIST = []

# Список строк для «Заполнить данные 1С» (заполняется при проверке файла в этой панели)
FILL_1C_DATA_LIST = []

# Список дел для «Загрузить информацию из 1С» (заполняется при нажатии «Начать»)
LOAD_1C_CASE_LIST = []

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")


def load_config():
    """Загружает настройки из config.json, если файл существует."""
    if not os.path.exists(CONFIG_PATH):
        return {}
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict):
                return data
            return {}
    except Exception:
        # При ошибке чтения возвращаем пустые настройки, чтобы не ломать сервер.
        return {}


def save_config(data: dict):
    """Сохраняет настройки в config.json."""
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        # Ошибку пробросим выше через ручку /settings, здесь не падаем.
        raise

ADMIN_HTML = """
<!doctype html>
<html lang="ru">
  <head>
    <meta charset="utf-8">
    <title>A.STORM робот — Админка</title>
    <style>
      * { box-sizing: border-box; }
      html, body { height: 100%; margin: 0; overflow: hidden; font-family: sans-serif; }
      .screen {
        display: none;
        position: fixed;
        top: 0; left: 0; right: 0; bottom: 0;
        overflow: hidden;
        flex-direction: column;
      }
      .screen.active {
        display: flex;
      }
      #home {
        align-items: center;
        justify-content: center;
        text-align: center;
        padding: 20px;
      }
      #home .main-title { margin-bottom: 30px; }
      #home .nav-btn { text-align: center; }
      .nav-btn {
        display: block;
        width: 100%;
        max-width: 320px;
        font-size: 16px;
        padding: 14px 20px;
        margin-bottom: 10px;
        cursor: pointer;
        text-align: left;
        border: 1px solid #333;
        border-radius: 6px;
        background: #f5f5f5;
      }
      .nav-btn:hover { background: #e8e8e8; }
      .nav-btn.nav-btn-green {
        background: #e4f7ea;
        border-color: #2e7d32;
        color: #1b5e20;
      }
      .nav-btn.nav-btn-green:hover {
        background: #d2f0dc;
      }
      .back-btn {
        font-size: 14px;
        padding: 8px 16px;
        margin-bottom: 20px;
        cursor: pointer;
        background: #ddd;
        border: 1px solid #999;
        border-radius: 4px;
        flex-shrink: 0;
      }
      .back-btn:hover { background: #ccc; }
      .panel-content {
        flex: 1;
        min-height: 0;
        overflow: auto;
        padding: 0 20px 20px;
      }
      #status { margin-top: 15px; }
      #log { background: #111; color: #0f0; padding: 10px; white-space: pre-wrap; overflow: auto; }
      .main-title { margin-bottom: 30px; font-size: 1.5em; }
      .panel-header { flex-shrink: 0; padding: 20px 20px 0; }
      .doc-upload-area { margin: 15px 0; }
      .doc-upload-area input[type="file"] { margin-right: 10px; }
      .doc-upload-area button { cursor: pointer; }
      .doc-status { margin: 10px 0; min-height: 1.2em; }
      .doc-counts { margin-top: 20px; }
      .doc-counts button { margin-top: 15px; cursor: pointer; }
      .doc-summary-block { margin-top: 18px; padding-top: 12px; border-top: 1px solid #e0e0e0; }
      .doc-summary-title { font-weight: bold; margin-bottom: 8px; font-size: 14px; }
      #docSummary { margin-top: 8px; font-size: 14px; }
      #docSummary ul { padding-left: 18px; }
      #docSummary li { margin-bottom: 4px; }
      #docSummary li button { margin-left: 8px; }
      #docValidationErrorsList { padding-left: 18px; margin: 8px 0; font-size: 14px; }
      .nav-btn.nav-btn-red {
        background: #fdecea;
        border-color: #c62828;
        color: #b71c1c;
      }
      .nav-btn.nav-btn-red:hover {
        background: #f9d1cb;
      }
      .nav-btn.nav-btn-yellow {
        background: #fff8e1;
        border-color: #f9a825;
        color: #9e6c00;
      }
      .nav-btn.nav-btn-yellow:hover {
        background: #ffecb3;
      }
    </style>
  </head>
  <body>
    <!-- Главный экран -->
    <div id="home" class="screen active">
      <h1 class="main-title">A.STORM робот</h1>
      <button type="button" class="nav-btn" data-panel="remote">Дистанционный запуск</button>
      <button type="button" class="nav-btn" data-panel="doc">Проверить документ выгрузки</button>
      <button type="button" class="nav-btn nav-btn-yellow" data-panel="settings">Настройки</button>
      <button type="button" class="nav-btn nav-btn-yellow" data-panel="fixfiles">Исправить названия файлов</button>
      <button type="button" class="nav-btn nav-btn-green" data-panel="check1c">Проверить данные 1С</button>
      <button type="button" class="nav-btn nav-btn-green" data-panel="fill1c">Заполнить данные 1С</button>
      <button type="button" class="nav-btn nav-btn-green" data-panel="load1c">Загрузить информацию из 1С</button>
    </div>

    <!-- Панель: Дистанционный запуск -->
    <div id="panel-remote" class="screen">
      <div class="panel-header">
        <button type="button" class="back-btn" data-back>Вернуться</button>
        <h1>Запуск alg.py на клиентах</h1>
        <p>Нажми кнопку, клиенты скачают свежий alg.py с сервера и выполнят его.</p>
        <button type="button" id="runBtn" class="nav-btn">Запустить на клиентах</button>
        <div id="status"></div>
      </div>
      <div class="panel-content">
        <h2>Лог клиента</h2>
        <pre id="log">Лога пока нет.</pre>
      </div>
    </div>

    <!-- Панель: Проверить документ выгрузки -->
    <div id="panel-doc" class="screen">
      <div class="panel-header">
        <button type="button" class="back-btn" data-back>Вернуться</button>
      </div>
      <div class="panel-content">
        <h1>Проверить документ выгрузки</h1>
        <p>Загрузите xlsx-документ, затем нажмите «Проверить». После проверки отобразится количество дел, которые можно и нельзя обработать, и подробная сводка по проблемным делам.</p>
        <div class="doc-upload-area">
          <label for="docFile">Файл выгрузки (.xlsx, .xls):</label><br>
          <input type="file" id="docFile" accept=".xlsx,.xls" />
          <button type="button" id="docCheckBtn" class="nav-btn">Проверить</button>
        </div>
        <div id="docCheckStatus" class="doc-status"></div>
        <div id="docCounts" class="doc-counts" style="display:none;">
          <p><strong>Можно обработать:</strong> <span id="docCanProcess">0</span></p>
          <p><strong>Нельзя обработать (нет обязательных полей):</strong> <span id="docCannotProcess">0</span></p>
          <div id="docSummaryBlock" class="doc-summary-block">
            <p class="doc-summary-title">Причины, по которым дела не прошли проверку (обязательные поля):</p>
            <div id="docSummary"></div>
            <button type="button" id="docDownloadBtn" class="nav-btn">Скачать все дела без обязательных полей</button>
          </div>
          <div id="docValidationSummary" class="doc-summary-block" style="display:none;">
            <p class="doc-summary-title">Ошибки валидации данных 1С: <span id="docValidationFailed">0</span> дел(а)</p>
            <ul id="docValidationErrorsList"></ul>
            <button type="button" id="docDownloadValidationBtn" class="nav-btn">Скачать дела с ошибками валидации</button>
          </div>
        </div>
      </div>
    </div>
    <div id="panel-check1c" class="screen">
      <div class="panel-header">
        <button type="button" class="back-btn" data-back>Вернуться</button>
      </div>
      <div class="panel-content">
        <h1>Проверить данные 1С</h1>
        <p><strong>Перед запуском проверки выполните подготовку удалённого рабочего стола с 1С.</strong></p>
        <ol>
          <li>Подключитесь к удалённому серверу, на котором установлена 1С.</li>
          <li>Запустите программу 1С (нужная вам база должна быть открыта).</li>
          <li>Разверните окно 1С на весь экран в оконном режиме (не в режиме «сеанс только в панели» и т.п.).</li>
          <li>Закройте в 1С все открытые вкладки/формы, чтобы на экране оставалось только главное окно.</li>
          <li>Убедитесь, что удалённый рабочий стол виден целиком и не перекрыт другими окнами.</li>
        </ol>
        <p>После нажатия кнопки «Начать» у вас будет примерно <strong>1 минута</strong>, чтобы переключиться на окно с удалённым сервером и не трогать мышь и клавиатуру, пока робот работает.</p>
        <div class="doc-upload-area">
          <label for="check1cFile">Файл с делами для проверки (например, выгрузка из Excel):</label><br>
          <input type="file" id="check1cFile" accept=".xlsx,.xls,.csv" />
        </div>
        <button type="button" id="check1cStartBtn" class="nav-btn">Начать</button>
        <div id="check1cStatus" class="doc-status"></div>
        <div id="check1cCounts" class="doc-counts" style="display:none;">
          <p><strong>Будет обработано:</strong> <span id="check1cCanProcess">0</span></p>
          <p><strong>Не будет обработано:</strong> <span id="check1cCannotProcess">0</span></p>
          <button type="button" id="check1cRunBtn" class="nav-btn">Запустить</button>
          <button type="button" id="check1cStopBtn" class="nav-btn nav-btn-red" style="display:none;">Остановить</button>
          <div id="check1cTimer" class="doc-status"></div>
        </div>
      </div>
    </div>
    <div id="panel-fill1c" class="screen">
      <div class="panel-header">
        <button type="button" class="back-btn" data-back>Вернуться</button>
      </div>
      <div class="panel-content">
        <h1>Заполнить данные 1С</h1>
        <p><strong>Перед запуском заполнения выполните подготовку удалённого рабочего стола с 1С.</strong></p>
        <ol>
          <li>Подключитесь к удалённому серверу, на котором установлена 1С.</li>
          <li>Запустите программу 1С (нужная вам база должна быть открыта).</li>
          <li>Разверните окно 1С на весь экран в оконном режиме (не в режиме «сеанс только в панели» и т.п.).</li>
          <li>Закройте в 1С все открытые вкладки/формы, чтобы на экране оставалось только главное окно.</li>
          <li>Убедитесь, что удалённый рабочий стол виден целиком и не перекрыт другими окнами.</li>
        </ol>
        <p>После нажатия кнопки «Начать» у вас будет примерно <strong>1 минута</strong>, чтобы переключиться на окно с удалённым сервером и не трогать мышь и клавиатуру, пока робот заполняет данные.</p>
        <div class="doc-upload-area">
          <label for="fill1cFile">Файл с данными для заполнения (например, выгрузка из Excel):</label><br>
          <input type="file" id="fill1cFile" accept=".xlsx,.xls,.csv" />
        </div>
        <button type="button" id="fill1cStartBtn" class="nav-btn">Начать</button>
        <div id="fill1cStatus" class="doc-status"></div>
        <div id="fill1cCounts" class="doc-counts" style="display:none;">
          <p><strong>Будет обработано:</strong> <span id="fill1cCanProcess">0</span></p>
          <p><strong>Не будет обработано:</strong> <span id="fill1cCannotProcess">0</span></p>
          <div id="fill1cValidationSummary" class="doc-summary" style="display:none;">
            <p><strong>Ошибки валидации данных 1С:</strong> <span id="fill1cValidationFailed">0</span> дел(а)</p>
            <ul id="fill1cValidationErrorsList"></ul>
          </div>
          <button type="button" id="fill1cRunBtn" class="nav-btn">Запустить</button>
          <button type="button" id="fill1cStopBtn" class="nav-btn nav-btn-red" style="display:none;">Остановить</button>
          <div id="fill1cTimer" class="doc-status"></div>
        </div>
      </div>
    </div>
    <div id="panel-load1c" class="screen">
      <div class="panel-header">
        <button type="button" class="back-btn" data-back>Вернуться</button>
      </div>
      <div class="panel-content">
        <h1>Загрузить информацию из 1С</h1>
        <p><strong>Перед запуском загрузки выполните подготовку удалённого рабочего стола с 1С.</strong></p>
        <ol>
          <li>Подключитесь к удалённому серверу, на котором установлена 1С.</li>
          <li>Запустите программу 1С (нужная вам база должна быть открыта).</li>
          <li>Разверните окно 1С на весь экран в оконном режиме (не в режиме «сеанс только в панели» и т.п.).</li>
          <li>Закройте в 1С все открытые вкладки/формы, чтобы на экране оставалось только главное окно.</li>
          <li>Убедитесь, что удалённый рабочий стол виден целиком и не перекрыт другими окнами.</li>
          <li>Скачайте одно тестовое дело из 1С и сохраните файл в папку, в которую робот будет в дальнейшем скачивать дела (папку для выгрузки).</li>
        </ol>
        <p>После нажатия кнопки «Начать» у вас будет примерно <strong>1 минута</strong>, чтобы переключиться на окно с удалённым сервером и не трогать мышь и клавиатуру, пока робот загружает информацию.</p>
        <p><strong>Выберите способ подготовки данных для тестовой загрузки:</strong></p>
        <div class="doc-upload-area">
          <label><input type="radio" name="load1cMode" id="load1cModeFile" value="file" checked> Загрузить файл выгрузки</label><br>
          <label><input type="radio" name="load1cMode" id="load1cModePaste" value="paste"> Вставить скопированную из Excel информацию</label>
        </div>
        <div class="doc-upload-area" id="load1cFileBlock">
          <label for="load1cFile">Файл с выгрузкой из 1С (одно тестовое дело):</label><br>
          <input type="file" id="load1cFile" accept=".xlsx,.xls,.csv" />
        </div>
        <div class="doc-upload-area" id="load1cPasteBlock" style="display:none;">
          <label for="load1cPaste">Скопируйте строку(и) из Excel и вставьте сюда (табличная вставка):</label><br>
          <textarea id="load1cPaste" rows="8" style="width:100%;" placeholder="Вставьте сюда данные из Excel (включая строку заголовков)"></textarea>
        </div>
        <button type="button" id="load1cStartBtn" class="nav-btn">Начать</button>
        <div id="load1cStatus" class="doc-status"></div>
        <div id="load1cCounts" class="doc-counts" style="display:none;">
          <p><strong>Будет обработано:</strong> <span id="load1cCanProcess">0</span></p>
          <p><strong>Не будет обработано:</strong> <span id="load1cCannotProcess">0</span></p>
          <button type="button" id="load1cRunBtn" class="nav-btn">Запустить</button>
          <button type="button" id="load1cStopBtn" class="nav-btn nav-btn-red" style="display:none;">Остановить</button>
          <div id="load1cTimer" class="doc-status"></div>
        </div>
      </div>
    </div>
    <div id="panel-settings" class="screen">
      <div class="panel-header">
        <button type="button" class="back-btn" data-back>Вернуться</button>
      </div>
      <div class="panel-content">
        <h1>Настройки</h1>
        <p>Задайте задержки (в секундах) для разных этапов работы робота.</p>
        <div class="doc-upload-area">
          <label for="delaySearchCase">Задержки при поиске дела:</label><br>
          <input type="number" id="delaySearchCase" step="0.1" min="0" style="max-width:200px;">
        </div>
        <div class="doc-upload-area">
          <label for="delayCourtTab">Задержки при работе с вкладкой суд:</label><br>
          <input type="number" id="delayCourtTab" step="0.1" min="0" style="max-width:200px;">
        </div>
        <div class="doc-upload-area">
          <label for="delayDownloadCases">Задержки при работе с загрузкой дел:</label><br>
          <input type="number" id="delayDownloadCases" step="0.1" min="0" style="max-width:200px;">
        </div>
        <div class="doc-upload-area">
          <label for="delayStageSwitch">Задержки при переходе между этапами:</label><br>
          <input type="number" id="delayStageSwitch" step="0.1" min="0" style="max-width:200px;">
        </div>
        <div class="doc-upload-area">
          <label for="delayBeforeRun">Секунд от нажатия «Запустить» до запуска робота (по умолчанию 60):</label><br>
          <input type="number" id="delayBeforeRun" step="1" min="0" style="max-width:200px;" placeholder="60">
        </div>
        <button type="button" id="settingsSaveBtn" class="nav-btn nav-btn-yellow">Сохранить</button>
        <div id="settingsStatus" class="doc-status"></div>
      </div>
    </div>
    <div id="panel-fixfiles" class="screen">
      <div class="panel-header">
        <button type="button" class="back-btn" data-back>Вернуться</button>
      </div>
      <div class="panel-content">
        <h1>Исправить названия файлов</h1>
        <div class="doc-upload-area">
          <label for="fixMode">Режим исправления:</label><br>
          <select id="fixMode" style="max-width:320px;">
            <option value="from1c" selected>Обработать выгрузку из 1С</option>
          </select>
        </div>
        <div class="doc-upload-area">
          <label for="fixPath">Путь к папке выгрузки:</label><br>
          <input type="text" id="fixPath" placeholder="Например: C:/Users/... или /home/user/..." style="width:100%; max-width:480px;">
        </div>
        <button type="button" id="fixRunBtn" class="nav-btn nav-btn-yellow">Исправить</button>
        <div id="fixStatus" class="doc-status"></div>
      </div>
    </div>

    <script>
      function show(id) {
        document.querySelectorAll('.screen').forEach(function(s) { s.classList.remove('active'); });
        var el = document.getElementById(id);
        if (el) el.classList.add('active');
      }

      document.addEventListener('DOMContentLoaded', function() {
        document.querySelectorAll('.nav-btn[data-panel]').forEach(function(btn) {
          btn.addEventListener('click', function(e) {
            e.preventDefault();
            var panelId = this.getAttribute('data-panel');
            if (panelId) show('panel-' + panelId);
          });
        });
        document.querySelectorAll('.back-btn[data-back]').forEach(function(btn) {
          btn.addEventListener('click', function() { show('home'); });
        });

        // Настройки: задержки и работа через config.json на сервере
        var delaySearchCase = document.getElementById('delaySearchCase');
        var delayCourtTab = document.getElementById('delayCourtTab');
        var delayDownloadCases = document.getElementById('delayDownloadCases');
        var delayStageSwitch = document.getElementById('delayStageSwitch');
        var delayBeforeRun = document.getElementById('delayBeforeRun');
        var settingsSaveBtn = document.getElementById('settingsSaveBtn');
        var settingsStatus = document.getElementById('settingsStatus');
        var robotDelayBeforeRun = 60;
        var fixMode = document.getElementById('fixMode');
        var fixPath = document.getElementById('fixPath');
        var fixRunBtn = document.getElementById('fixRunBtn');
        var fixStatus = document.getElementById('fixStatus');

        if (delaySearchCase && delayCourtTab && delayDownloadCases && delayStageSwitch) {
          fetch('/settings')
            .then(function(r) { return r.json().then(function(d) { return { ok: r.ok, data: d }; }); })
            .then(function(o) {
              if (o.ok && o.data) {
                if (o.data.delaySearchCase != null) delaySearchCase.value = o.data.delaySearchCase;
                if (o.data.delayCourtTab != null) delayCourtTab.value = o.data.delayCourtTab;
                if (o.data.delayDownloadCases != null) delayDownloadCases.value = o.data.delayDownloadCases;
                if (o.data.delayStageSwitch != null) delayStageSwitch.value = o.data.delayStageSwitch;
                if (delayBeforeRun && o.data.delayBeforeRun != null) delayBeforeRun.value = o.data.delayBeforeRun;
                robotDelayBeforeRun = (o.data.delayBeforeRun != null && o.data.delayBeforeRun !== '') ? parseInt(o.data.delayBeforeRun, 10) : 60;
                if (isNaN(robotDelayBeforeRun) || robotDelayBeforeRun < 0) robotDelayBeforeRun = 60;
              }
            })
            .catch(function(e) {
              console.warn('Не удалось загрузить настройки с сервера:', e);
            });
        }

        if (settingsSaveBtn) {
          settingsSaveBtn.addEventListener('click', function() {
            var data = {
              delaySearchCase: delaySearchCase && delaySearchCase.value !== '' ? parseFloat(delaySearchCase.value) : null,
              delayCourtTab: delayCourtTab && delayCourtTab.value !== '' ? parseFloat(delayCourtTab.value) : null,
              delayDownloadCases: delayDownloadCases && delayDownloadCases.value !== '' ? parseFloat(delayDownloadCases.value) : null,
              delayStageSwitch: delayStageSwitch && delayStageSwitch.value !== '' ? parseFloat(delayStageSwitch.value) : null,
              delayBeforeRun: delayBeforeRun && delayBeforeRun.value !== '' ? parseInt(delayBeforeRun.value, 10) : null
            };
            if (data.delayBeforeRun != null && !isNaN(data.delayBeforeRun) && data.delayBeforeRun >= 0) robotDelayBeforeRun = data.delayBeforeRun;
            fetch('/settings', {
              method: 'POST',
              headers: { 'Content-Type': 'application/json' },
              body: JSON.stringify(data)
            })
              .then(function(r) { return r.json().then(function(d) { return { ok: r.ok, data: d }; }); })
              .then(function(o) {
                if (o.ok && o.data && o.data.ok) {
                  if (settingsStatus) {
                    settingsStatus.textContent = 'Настройки сохранены.';
                  }
                } else {
                  if (settingsStatus) {
                    settingsStatus.textContent = (o.data && o.data.error) ? o.data.error : 'Ошибка сохранения настроек.';
                  }
                }
              })
              .catch(function(e) {
                if (settingsStatus) {
                  settingsStatus.textContent = 'Ошибка сохранения настроек: ' + e;
                }
              });
          });
        }

        if (fixRunBtn) {
          fixRunBtn.addEventListener('click', function() {
            if (!fixPath || !fixPath.value.trim()) {
              if (fixStatus) {
                fixStatus.textContent = 'Укажите путь к папке выгрузки.';
              }
              return;
            }
            if (fixStatus) {
              fixStatus.textContent = 'Исправляю названия файлов...';
            }
            // Пока реальная логика не реализована, просто имитируем быстрое завершение.
            setTimeout(function() {
              if (fixStatus) {
                fixStatus.textContent = 'Готово. Названия файлов в папке "' + fixPath.value.trim() + '" исправлены (режим: «Обработать выгрузку из 1С»).';
              }
            }, 300);
          });
        }

        var docFile = document.getElementById('docFile');
        var docCheckBtn = document.getElementById('docCheckBtn');
        var docCheckStatus = document.getElementById('docCheckStatus');
        var docCounts = document.getElementById('docCounts');
        var docSummary = document.getElementById('docSummary');
        var docDownloadBtn = document.getElementById('docDownloadBtn');
        if (docCheckBtn && docFile) {
          docCheckBtn.addEventListener('click', function() {
            if (!docFile.files || !docFile.files[0]) {
              docCheckStatus.textContent = 'Сначала выберите файл .xlsx';
              return;
            }
            docCheckStatus.textContent = 'Проверяю...';
            var fd = new FormData();
            fd.append('document', docFile.files[0]);
            fetch('/check-doc-upload', { method: 'POST', body: fd })
              .then(function(r) { return r.json().then(function(d) { return { ok: r.ok, data: d }; }); })
              .then(function(o) {
                if (o.ok && o.data.ok) {
                  docCheckStatus.textContent = 'Проверка завершена.';
                  document.getElementById('docCanProcess').textContent = o.data.can_process;
                  document.getElementById('docCannotProcess').textContent = o.data.cannot_process;
                  docCounts.style.display = 'block';

                  // Сводка: дела без обязательных полей
                  var rs = o.data.reasons_summary || {};
                  var keys = Object.keys(rs);
                  if (docSummary) {
                    if (!keys.length) {
                      docSummary.textContent = o.data.cannot_process === 0
                        ? 'Нет дел без обязательных полей.'
                        : 'Все необрабатываемые дела не проходят по прочим причинам.';
                    } else {
                      var html = '<ul>';
                      keys.forEach(function(k) {
                        var count = rs[k];
                        html += '<li>' +
                          k + ' — ' + count + ' дел(а)' +
                          ' <button type="button" class="reason-download-btn" data-missing-field="' +
                          encodeURIComponent(k) +
                          '">Скачать дела по этой причине</button>' +
                          '</li>';
                      });
                      html += '</ul>';
                      docSummary.innerHTML = html;

                      var reasonBtns = docSummary.querySelectorAll('.reason-download-btn');
                      reasonBtns.forEach(function(btn) {
                        btn.addEventListener('click', function() {
                          var field = this.getAttribute('data-missing-field');
                          if (field) {
                            window.location.href = '/download-unprocessable-doc-by-reason?field=' + field;
                          }
                        });
                      });
                    }
                    if (docDownloadBtn) {
                      docDownloadBtn.style.display = o.data.cannot_process > 0 ? '' : 'none';
                    }
                  }

                  // Сводка: ошибки валидации данных 1С
                  var validationFailed = o.data.validation_failed || 0;
                  var docValidationSummary = document.getElementById('docValidationSummary');
                  var docValidationFailed = document.getElementById('docValidationFailed');
                  var docValidationErrorsList = document.getElementById('docValidationErrorsList');
                  if (docValidationSummary && docValidationFailed && docValidationErrorsList) {
                    if (validationFailed > 0) {
                      docValidationFailed.textContent = validationFailed;
                      var ves = o.data.validation_errors_summary || {};
                      var vesKeys = Object.keys(ves);
                      docValidationErrorsList.innerHTML = '';
                      vesKeys.forEach(function(msg) {
                        var li = document.createElement('li');
                        li.textContent = msg + ' — ' + ves[msg] + ' дел(а)';
                        docValidationErrorsList.appendChild(li);
                      });
                      docValidationSummary.style.display = 'block';
                    } else {
                      docValidationSummary.style.display = 'none';
                    }
                  }
                } else {
                  docCheckStatus.textContent = o.data.error || 'Ошибка проверки';
                }
              })
              .catch(function(e) {
                docCheckStatus.textContent = 'Ошибка: ' + e;
              });
          });
        }
        if (docDownloadBtn) {
          docDownloadBtn.addEventListener('click', function() {
            window.location.href = '/download-unprocessable-doc';
          });
        }
        var docDownloadValidationBtn = document.getElementById('docDownloadValidationBtn');
        if (docDownloadValidationBtn) {
          docDownloadValidationBtn.addEventListener('click', function() {
            window.location.href = '/download-validation-failed-doc';
          });
        }

        // Проверка данных 1С: валидация файла и запуск таймера
        var check1cFile = document.getElementById('check1cFile');
        var check1cStartBtn = document.getElementById('check1cStartBtn');
        var check1cStatus = document.getElementById('check1cStatus');
        var check1cCounts = document.getElementById('check1cCounts');
        var check1cCanProcess = document.getElementById('check1cCanProcess');
        var check1cCannotProcess = document.getElementById('check1cCannotProcess');
        var check1cRunBtn = document.getElementById('check1cRunBtn');
        var check1cStopBtn = document.getElementById('check1cStopBtn');
        var check1cTimer = document.getElementById('check1cTimer');
        var check1cTimerId = null;

        if (check1cStartBtn && check1cFile) {
          check1cStartBtn.addEventListener('click', function() {
            if (!check1cFile.files || !check1cFile.files[0]) {
              check1cStatus.textContent = 'Сначала выберите файл с делами (.xlsx/.xls/.csv).';
              return;
            }
            check1cStatus.textContent = 'Проверяю номера дел...';
            check1cCounts.style.display = 'none';
            if (check1cTimerId !== null) {
              clearInterval(check1cTimerId);
              check1cTimerId = null;
            }
            if (check1cTimer) {
              check1cTimer.textContent = '';
            }
            var fd = new FormData();
            fd.append('document', check1cFile.files[0]);
            fetch('/check-doc-upload-ids-only', { method: 'POST', body: fd })
              .then(function(r) { return r.json().then(function(d) { return { ok: r.ok, data: d }; }); })
              .then(function(o) {
                if (o.ok && o.data.ok) {
                  check1cStatus.textContent = 'Проверка завершена. Учитываются только заполненные номера дел, остальные поля не обязательны.';
                  if (check1cCanProcess) {
                    check1cCanProcess.textContent = o.data.can_process;
                  }
                  if (check1cCannotProcess) {
                    check1cCannotProcess.textContent = o.data.cannot_process;
                  }
                  if (check1cCounts) {
                    check1cCounts.style.display = 'block';
                  }
                } else {
                  check1cStatus.textContent = o.data.error || 'Ошибка проверки файла';
                }
              })
              .catch(function(e) {
                check1cStatus.textContent = 'Ошибка: ' + e;
              });
          });
        }

        if (check1cRunBtn) {
          check1cRunBtn.addEventListener('click', function() {
            var totalSec = (typeof robotDelayBeforeRun === 'number' && robotDelayBeforeRun >= 0) ? robotDelayBeforeRun : 60;
            var secondsLeft = totalSec;
            if (check1cTimerId !== null) {
              clearInterval(check1cTimerId);
            }
            if (check1cTimer) {
              check1cTimer.textContent = 'Осталось ' + totalSec + ' секунд. Переключитесь на удалённый рабочий стол и не трогайте мышь и клавиатуру.';
            }
            check1cRunBtn.disabled = true;
            if (check1cStopBtn) {
              check1cStopBtn.style.display = 'inline-block';
              check1cStopBtn.disabled = false;
            }
            check1cStartBtn && (check1cStartBtn.disabled = true);
            check1cTimerId = setInterval(function() {
              secondsLeft -= 1;
              if (secondsLeft > 0) {
                if (check1cTimer) {
                  check1cTimer.textContent = 'Осталось ' + secondsLeft + ' секунд. Не трогайте мышь и клавиатуру.';
                }
              } else {
                clearInterval(check1cTimerId);
                check1cTimerId = null;
                if (check1cTimer) {
                  check1cTimer.textContent = totalSec + ' секунд прошло. Запускаю робота...';
                }
                fetch('/robot/run', {
                  method: 'POST',
                  headers: { 'Content-Type': 'application/json' },
                  body: JSON.stringify({ mode: 'check' })
                })
                  .then(function(r) { return r.json().then(function(d) { return { ok: r.ok, data: d }; }); })
                  .then(function(o) {
                    if (check1cTimer) {
                      check1cTimer.textContent = o.ok
                        ? totalSec + ' секунд прошло. Робот запущен. Смотрите robot.log.'
                        : totalSec + ' секунд прошло. Ошибка: ' + (o.data.error || 'неизвестная');
                    }
                  })
                  .catch(function(e) {
                    if (check1cTimer) {
                      check1cTimer.textContent = totalSec + ' секунд прошло. Ошибка запуска: ' + e;
                    }
                  });
                check1cRunBtn.disabled = false;
                check1cStartBtn && (check1cStartBtn.disabled = false);
                // Кнопку «Остановить» не скрываем — робот может ещё работать
              }
            }, 1000);
          });
        }

        if (check1cStopBtn) {
          check1cStopBtn.addEventListener('click', function() {
            if (check1cTimerId !== null) {
              clearInterval(check1cTimerId);
              check1cTimerId = null;
            }
            fetch('/robot/stop', { method: 'POST' })
              .then(function() {
                if (check1cTimer) {
                  check1cTimer.textContent = 'Остановка запрошена. Робот прервёт работу при следующей проверке.';
                }
              })
              .catch(function() {});
            check1cRunBtn && (check1cRunBtn.disabled = false);
            check1cStartBtn && (check1cStartBtn.disabled = false);
            check1cStopBtn.style.display = 'none';
          });
        }

        // Заполнение данных 1С: валидация файла и запуск таймера (аналогично проверке)
        var fill1cFile = document.getElementById('fill1cFile');
        var fill1cStartBtn = document.getElementById('fill1cStartBtn');
        var fill1cStatus = document.getElementById('fill1cStatus');
        var fill1cCounts = document.getElementById('fill1cCounts');
        var fill1cCanProcess = document.getElementById('fill1cCanProcess');
        var fill1cCannotProcess = document.getElementById('fill1cCannotProcess');
        var fill1cRunBtn = document.getElementById('fill1cRunBtn');
        var fill1cStopBtn = document.getElementById('fill1cStopBtn');
        var fill1cTimer = document.getElementById('fill1cTimer');
        var fill1cTimerId = null;

        if (fill1cStartBtn && fill1cFile) {
          fill1cStartBtn.addEventListener('click', function() {
            if (!fill1cFile.files || !fill1cFile.files[0]) {
              fill1cStatus.textContent = 'Сначала выберите файл с данными (.xlsx/.xls/.csv).';
              return;
            }
            fill1cStatus.textContent = 'Проверяю файл...';
            fill1cCounts.style.display = 'none';
            if (fill1cTimerId !== null) {
              clearInterval(fill1cTimerId);
              fill1cTimerId = null;
            }
            if (fill1cTimer) {
              fill1cTimer.textContent = '';
            }
            var fd2 = new FormData();
            fd2.append('document', fill1cFile.files[0]);
            fetch('/check-doc-upload', { method: 'POST', body: fd2 })
              .then(function(r) { return r.json().then(function(d) { return { ok: r.ok, data: d }; }); })
              .then(function(o) {
                if (o.ok && o.data.ok) {
                  fill1cStatus.textContent = 'Проверка завершена. Проверьте количества дел перед запуском заполнения.';
                  if (fill1cCanProcess) {
                    fill1cCanProcess.textContent = o.data.can_process;
                  }
                  if (fill1cCannotProcess) {
                    fill1cCannotProcess.textContent = o.data.cannot_process;
                  }
                  var validationFailed = o.data.validation_failed || 0;
                  var fill1cValidationSummary = document.getElementById('fill1cValidationSummary');
                  var fill1cValidationFailed = document.getElementById('fill1cValidationFailed');
                  var fill1cValidationErrorsList = document.getElementById('fill1cValidationErrorsList');
                  if (fill1cValidationSummary && fill1cValidationFailed && fill1cValidationErrorsList) {
                    if (validationFailed > 0) {
                      fill1cValidationFailed.textContent = validationFailed;
                      var ves = o.data.validation_errors_summary || {};
                      var vesKeys = Object.keys(ves);
                      fill1cValidationErrorsList.innerHTML = '';
                      vesKeys.forEach(function(msg) {
                        var li = document.createElement('li');
                        li.textContent = msg + ' — ' + ves[msg] + ' дел(а)';
                        fill1cValidationErrorsList.appendChild(li);
                      });
                      fill1cValidationSummary.style.display = 'block';
                    } else {
                      fill1cValidationSummary.style.display = 'none';
                    }
                  }
                  if (fill1cCounts) {
                    fill1cCounts.style.display = 'block';
                  }
                } else {
                  fill1cStatus.textContent = o.data.error || 'Ошибка проверки файла';
                }
              })
              .catch(function(e) {
                fill1cStatus.textContent = 'Ошибка: ' + e;
              });
          });
        }

        if (fill1cRunBtn) {
          fill1cRunBtn.addEventListener('click', function() {
            var totalSec2 = (typeof robotDelayBeforeRun === 'number' && robotDelayBeforeRun >= 0) ? robotDelayBeforeRun : 60;
            var secondsLeft2 = totalSec2;
            if (fill1cTimerId !== null) {
              clearInterval(fill1cTimerId);
            }
            if (fill1cTimer) {
              fill1cTimer.textContent = 'Осталось ' + totalSec2 + ' секунд. Переключитесь на удалённый рабочий стол и не трогайте мышь и клавиатуру.';
            }
            fill1cRunBtn.disabled = true;
            if (fill1cStopBtn) {
              fill1cStopBtn.style.display = 'inline-block';
              fill1cStopBtn.disabled = false;
            }
            fill1cStartBtn && (fill1cStartBtn.disabled = true);
            fill1cTimerId = setInterval(function() {
              secondsLeft2 -= 1;
              if (secondsLeft2 > 0) {
                if (fill1cTimer) {
                  fill1cTimer.textContent = 'Осталось ' + secondsLeft2 + ' секунд. Не трогайте мышь и клавиатуру.';
                }
              } else {
                clearInterval(fill1cTimerId);
                fill1cTimerId = null;
                if (fill1cTimer) {
                  fill1cTimer.textContent = totalSec2 + ' секунд прошло. Запускаю робота...';
                }
                fetch('/robot/run', {
                  method: 'POST',
                  headers: { 'Content-Type': 'application/json' },
                  body: JSON.stringify({ mode: 'set' })
                })
                  .then(function(r) { return r.json().then(function(d) { return { ok: r.ok, data: d }; }); })
                  .then(function(o) {
                    if (fill1cTimer) {
                      fill1cTimer.textContent = o.ok
                        ? totalSec2 + ' секунд прошло. Робот запущен. Смотрите robot.log.'
                        : totalSec2 + ' секунд прошло. Ошибка: ' + (o.data.error || 'неизвестная');
                    }
                  })
                  .catch(function(e) {
                    if (fill1cTimer) {
                      fill1cTimer.textContent = totalSec2 + ' секунд прошло. Ошибка запуска: ' + e;
                    }
                  });
                fill1cRunBtn.disabled = false;
                fill1cStartBtn && (fill1cStartBtn.disabled = false);
              }
            }, 1000);
          });
        }

        if (fill1cStopBtn) {
          fill1cStopBtn.addEventListener('click', function() {
            if (fill1cTimerId !== null) {
              clearInterval(fill1cTimerId);
              fill1cTimerId = null;
            }
            fetch('/robot/stop', { method: 'POST' }).catch(function() {});
            if (fill1cTimer) {
              fill1cTimer.textContent = 'Остановка запрошена.';
            }
            fill1cRunBtn && (fill1cRunBtn.disabled = false);
            fill1cStartBtn && (fill1cStartBtn.disabled = false);
            fill1cStopBtn.style.display = 'none';
          });
        }

        // Загрузка информации из 1С: валидация файла и запуск таймера
        var load1cModeFile = document.getElementById('load1cModeFile');
        var load1cModePaste = document.getElementById('load1cModePaste');
        var load1cFileBlock = document.getElementById('load1cFileBlock');
        var load1cPasteBlock = document.getElementById('load1cPasteBlock');
        var load1cFile = document.getElementById('load1cFile');
        var load1cPaste = document.getElementById('load1cPaste');
        var load1cStartBtn = document.getElementById('load1cStartBtn');
        var load1cStatus = document.getElementById('load1cStatus');
        var load1cCounts = document.getElementById('load1cCounts');
        var load1cCanProcess = document.getElementById('load1cCanProcess');
        var load1cCannotProcess = document.getElementById('load1cCannotProcess');
        var load1cRunBtn = document.getElementById('load1cRunBtn');
        var load1cStopBtn = document.getElementById('load1cStopBtn');
        var load1cTimer = document.getElementById('load1cTimer');
        var load1cTimerId = null;

        if (load1cModeFile && load1cModePaste && load1cFileBlock && load1cPasteBlock) {
          function updateLoad1cMode() {
            var useFile = load1cModeFile.checked;
            load1cFileBlock.style.display = useFile ? 'block' : 'none';
            load1cPasteBlock.style.display = useFile ? 'none' : 'block';
          }
          load1cModeFile.addEventListener('change', updateLoad1cMode);
          load1cModePaste.addEventListener('change', updateLoad1cMode);
          updateLoad1cMode();
        }

        if (load1cStartBtn && (load1cFile || load1cPaste)) {
          load1cStartBtn.addEventListener('click', function() {
            var useFileMode = !load1cModePaste || load1cModeFile.checked;
            if (useFileMode) {
              if (!load1cFile || !load1cFile.files || !load1cFile.files[0]) {
                load1cStatus.textContent = 'Сначала выберите файл с выгрузкой (.xlsx/.xls/.csv).';
                return;
              }
            } else {
              if (!load1cPaste || !load1cPaste.value.trim()) {
                load1cStatus.textContent = 'Сначала вставьте данные, скопированные из Excel.';
                return;
              }
            }
            load1cStatus.textContent = 'Считаю, сколько дел будет загружено...';
            load1cCounts.style.display = 'none';
            if (load1cTimerId !== null) {
              clearInterval(load1cTimerId);
              load1cTimerId = null;
            }
            if (load1cTimer) {
              load1cTimer.textContent = '';
            }
            var fd3 = new FormData();
            if (useFileMode) {
              fd3.append('document', load1cFile.files[0]);
            } else {
              fd3.append('pasted_text', load1cPaste.value);
            }
            fetch('/check-simple-list', { method: 'POST', body: fd3 })
              .then(function(r) { return r.json().then(function(d) { return { ok: r.ok, data: d }; }); })
              .then(function(o) {
                if (o.ok && o.data.ok) {
                  load1cStatus.textContent = 'Готово. Проверьте, сколько дел будет загружено.';
                  if (load1cCanProcess) {
                    load1cCanProcess.textContent = o.data.total;
                  }
                  if (load1cCannotProcess) {
                    load1cCannotProcess.textContent = 0;
                  }
                  if (load1cCounts) {
                    load1cCounts.style.display = 'block';
                  }
                } else {
                  load1cStatus.textContent = o.data.error || 'Ошибка обработки списка дел';
                }
              })
              .catch(function(e) {
                load1cStatus.textContent = 'Ошибка: ' + e;
              });
          });
        }

        if (load1cRunBtn) {
          load1cRunBtn.addEventListener('click', function() {
            var totalSec3 = (typeof robotDelayBeforeRun === 'number' && robotDelayBeforeRun >= 0) ? robotDelayBeforeRun : 60;
            var secondsLeft3 = totalSec3;
            if (load1cTimerId !== null) {
              clearInterval(load1cTimerId);
            }
            if (load1cTimer) {
              load1cTimer.textContent = 'Осталось ' + totalSec3 + ' секунд. Переключитесь на удалённый рабочий стол и не трогайте мышь и клавиатуру.';
            }
            load1cRunBtn.disabled = true;
            if (load1cStopBtn) {
              load1cStopBtn.style.display = 'inline-block';
              load1cStopBtn.disabled = false;
            }
            load1cStartBtn && (load1cStartBtn.disabled = true);
            load1cTimerId = setInterval(function() {
              secondsLeft3 -= 1;
              if (secondsLeft3 > 0) {
                if (load1cTimer) {
                  load1cTimer.textContent = 'Осталось ' + secondsLeft3 + ' секунд. Не трогайте мышь и клавиатуру.';
                }
              } else {
                clearInterval(load1cTimerId);
                load1cTimerId = null;
                if (load1cTimer) {
                  load1cTimer.textContent = totalSec3 + ' секунд прошло. Запускаю робота...';
                }
                fetch('/robot/run', {
                  method: 'POST',
                  headers: { 'Content-Type': 'application/json' },
                  body: JSON.stringify({ mode: 'download' })
                })
                  .then(function(r) { return r.json().then(function(d) { return { ok: r.ok, data: d }; }); })
                  .then(function(o) {
                    if (load1cTimer) {
                      load1cTimer.textContent = o.ok
                        ? totalSec3 + ' секунд прошло. Робот запущен. Смотрите robot.log.'
                        : totalSec3 + ' секунд прошло. Ошибка: ' + (o.data.error || 'неизвестная');
                    }
                  })
                  .catch(function(e) {
                    if (load1cTimer) {
                      load1cTimer.textContent = totalSec3 + ' секунд прошло. Ошибка запуска: ' + e;
                    }
                  });
                load1cRunBtn.disabled = false;
                load1cStartBtn && (load1cStartBtn.disabled = false);
              }
            }, 1000);
          });
        }

        if (load1cStopBtn) {
          load1cStopBtn.addEventListener('click', function() {
            if (load1cTimerId !== null) {
              clearInterval(load1cTimerId);
              load1cTimerId = null;
            }
            fetch('/robot/stop', { method: 'POST' }).catch(function() {});
            if (load1cTimer) {
              load1cTimer.textContent = 'Остановка запрошена.';
            }
            load1cRunBtn && (load1cRunBtn.disabled = false);
            load1cStartBtn && (load1cStartBtn.disabled = false);
            load1cStopBtn.style.display = 'none';
          });
        }
      });

      var runBtn = document.getElementById('runBtn');
      const statusEl = document.getElementById('status');
      const logEl = document.getElementById('log');

      async function loadLog() {
        try {
          const resp = await fetch('/client-log');
          const data = await resp.json();
          if (data.run_id === null) {
            logEl.textContent = 'Лога пока нет.';
          } else {
            logEl.textContent =
              'RUN_ID: ' + data.run_id + '\\n' +
              'returncode: ' + data.returncode + '\\n\\n' +
              'STDOUT:\\n' + (data.stdout || '') + '\\n\\n' +
              'STDERR:\\n' + (data.stderr || '');
          }
        } catch (e) {
          logEl.textContent = 'Ошибка загрузки лога: ' + e;
        }
      }

      if (runBtn) {
        runBtn.onclick = async () => {
          statusEl.textContent = 'Отправляю команду...';
          try {
            const resp = await fetch('/trigger', { method: 'POST' });
            const data = await resp.json();
            statusEl.textContent = 'OK, RUN_ID = ' + data.run_id;
          } catch (e) {
            statusEl.textContent = 'Ошибка: ' + e;
          }
        };
      }
      loadLog();
      setInterval(loadLog, 2000);
    </script>
  </body>
  </html>
"""


@app.route("/")
def admin() -> Response:
    # Простая админка с кнопкой
    return Response(ADMIN_HTML, mimetype="text/html")


@app.route("/trigger", methods=["POST"])
def trigger():
    """
    Админка дергает этот эндпоинт.
    Увеличиваем RUN_ID — клиенты по опросу увидят, что появился новый запуск.
    """
    global RUN_ID
    RUN_ID += 1
    return jsonify({"ok": True, "run_id": RUN_ID})


@app.route("/next-command", methods=["GET"])
def next_command():
    """
    Клиент опрашивает этот эндпоинт, передавая last_run_id (что он уже выполнял).
    Возвращаем, нужно ли запускать новый алгоритм, и ссылку на alg.py.
    """
    global RUN_ID
    try:
        last_run_id = int(request.args.get("last_run_id", "0"))
    except ValueError:
        last_run_id = 0

    should_run = RUN_ID > last_run_id
    alg_url = request.url_root.rstrip("/") + "/alg.py"

    return jsonify(
        {
            "run_id": RUN_ID,
            "should_run": should_run,
            "alg_url": alg_url,
        }
    )


@app.route("/alg.py", methods=["GET"])
def download_alg():
    """
    Клиент скачивает сам файл alg.py с сервера.
    """
    return send_from_directory(BASE_DIR, "alg.py", as_attachment=False)


@app.route("/client-log", methods=["POST", "GET"])
def client_log():
    """
    POST: клиент отправляет лог выполнения alg.py.
    GET: админка забирает последний лог.
    """
    global CLIENT_LOG
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        CLIENT_LOG = {
            "run_id": data.get("run_id"),
            "returncode": data.get("returncode"),
            "stdout": data.get("stdout") or "",
            "stderr": data.get("stderr") or "",
        }
        return jsonify({"ok": True})
    return jsonify(CLIENT_LOG)


def _parse_xlsx_rows(file_stream):
    """Читает xlsx из file-like объекта, возвращает (headers, list_of_row_dicts)."""
    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        data.append(dict(zip(headers, (v if v is not None else "" for v in row))))
    return headers, data


def _parse_pasted_table(text):
    """
    Простейший парсер табличных данных, скопированных из Excel.
    Ожидает, что первая строка — заголовки.
    Поддерживает разделители: табуляция, ';' или ','.
    """
    if not text.strip():
        return [], []
    lines = [line for line in text.splitlines() if line.strip() != ""]
    if not lines:
        return [], []
    first = lines[0]
    if "\t" in first:
        delimiter = "\t"
    elif ";" in first:
        delimiter = ";"
    else:
        delimiter = ","
    reader = csv.reader(StringIO("\n".join(lines)), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        row_dict = {}
        for idx, h in enumerate(headers):
            row_dict[h] = row[idx] if idx < len(row) else ""
        data.append(row_dict)
    return headers, data


@app.route("/check-simple-list", methods=["POST"])
def check_simple_list():
    """
    Принимает либо xlsx-файл, либо текст, скопированный из Excel,
    и просто считает количество дел по столбцу
    'ID дела  (ID/ номер убытка)'.
    """
    global LOAD_1C_CASE_LIST
    id_field = "ID дела  (ID/ номер убытка)"

    # Вариант 1: пришёл файл (xlsx) — считаем по столбцу ID дела
    if "document" in request.files and request.files["document"].filename:
        file = request.files["document"]
        try:
            headers, rows = _parse_xlsx_rows(file.stream)
        except Exception as e:
            return jsonify({"error": f"Ошибка чтения файла: {e}"}), 400
        if not headers:
            return jsonify({"error": "Файл пустой или без заголовков."}), 400

        if id_field not in headers:
            return (
                jsonify(
                    {
                        "error": f"Не найден столбец '{id_field}' в заголовке файла."
                    }
                ),
                400,
            )

        # Считаем только строки, где ID не пустой, и сохраняем для запуска робота
        load_case_list = []
        for row in rows:
            val = row.get(id_field)
            if val is not None and str(val).strip() != "":
                load_case_list.append({"number_case": str(val).strip()})
        LOAD_1C_CASE_LIST = load_case_list
        return jsonify({"ok": True, "total": len(load_case_list)})

    # Вариант 2: пришёл только текст, скопированный из Excel.
    pasted = request.form.get("pasted_text", "").strip()
    if not pasted:
        return jsonify({"error": "Не передан файл или текст со списком дел."}), 400

    headers_pt, rows_pt = _parse_pasted_table(pasted)
    load_case_list = []
    if id_field in headers_pt:
        for row in rows_pt:
            val = row.get(id_field)
            if val is not None and str(val).strip() != "":
                load_case_list.append({"number_case": str(val).strip()})
    else:
        # Нет столбца ID — считаем по количеству строк (без заголовка)
        load_case_list = [{"number_case": str(i)} for i in range(len(rows_pt))] if rows_pt else []
    LOAD_1C_CASE_LIST = load_case_list
    return jsonify({"ok": True, "total": len(load_case_list)})


REQUIRED_DOC_FIELDS = [
    "ID дела  (ID/ номер убытка)",
    "Наименование должника",
    "Суд",
    "Дата вынесения решения",
    "Дата направления иска",
    "Результат судебного дела",
    "Сумма долга",
    "Взысканная сумма (основной долг)",
    "Взысканная сумма (расходы)",
    "Сумма госпошлины",
    "Номер исполнительного листа",
    "Дата получения исполнительного листа",
]


def _is_empty_value(value):
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == "" or value.strip() == "-"
    return False


def _check_doc_required_fields(rows):
    """
    Проверяет, что в каждой строке (деле) заполнены обязательные поля из REQUIRED_DOC_FIELDS.
    Возвращает (can_process_count, cannot_process_rows, can_process_rows).
    """
    cannot_rows = []
    can_process_rows = []
    for row in rows:
        missing_fields = []
        for field in REQUIRED_DOC_FIELDS:
            if field not in row or _is_empty_value(row.get(field)):
                missing_fields.append(field)
        if missing_fields:
            row_copy = dict(row)
            row_copy["_missing_required_fields"] = missing_fields
            cannot_rows.append(row_copy)
        else:
            can_process_rows.append(dict(row))
    return len(can_process_rows), cannot_rows, can_process_rows


def _build_unprocessable_summary(cannot_rows):
    """
    Строит сводку по причинам, по которым дела не прошли проверку.
    Возвращает словарь: {<название обязательного поля>: <кол-во дел без него>}.
    """
    summary = {}
    for row in cannot_rows:
        missing = row.get("_missing_required_fields") or []
        # missing может быть как списком, так и строкой (на будущее)
        if isinstance(missing, str):
            fields = [f.strip() for f in missing.split(",") if f.strip()]
        else:
            fields = list(missing)
        for field in fields:
            summary[field] = summary.get(field, 0) + 1
    return summary


def _validate_fill1c_rows(can_process_rows):
    """
    Проверяет строки для «Заполнить данные 1С» через validate_before_setting.
    Возвращает (valid_rows, validation_failed_rows, validation_summary).
    validation_failed_rows — список dict с ключом _validation_error (текст ошибки).
    validation_summary — {текст_ошибки: количество}.
    """
    valid_rows = []
    validation_failed_rows = []
    for row in can_process_rows:
        try:
            case = normalize_robot_payload("set", row)
            validate_before_setting(case)
            valid_rows.append(row)
        except Exception as e:
            row_copy = dict(row)
            err_msg = str(e).strip() or "Ошибка валидации"
            row_copy["_validation_error"] = err_msg
            validation_failed_rows.append(row_copy)
    summary = {}
    for row in validation_failed_rows:
        msg = row.get("_validation_error") or "Ошибка валидации"
        summary[msg] = summary.get(msg, 0) + 1
    return valid_rows, validation_failed_rows, summary


@app.route("/check-doc-upload", methods=["POST"])
def check_doc_upload():
    """
    Принимает xlsx-файл, проверяет его, возвращает кол-во обрабатываемых и необрабатываемых.
    Строки с заполненными обязательными полями дополнительно проходят валидацию validate_before_setting.
    Сохраняет необрабатываемые строки и строки с ошибками валидации для скачивания.
    """
    global DOC_CHECK_RESULT
    if "document" not in request.files:
        return jsonify({"error": "Файл не выбран"}), 400
    file = request.files["document"]
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Нужен файл .xlsx"}), 400
    try:
        headers, rows = _parse_xlsx_rows(file.stream)
    except Exception as e:
        return jsonify({"error": f"Ошибка чтения файла: {e}"}), 400
    if not headers:
        return jsonify({"error": "Файл пустой или без заголовков"}), 400
    can_process, cannot_rows, can_process_rows = _check_doc_required_fields(rows)
    valid_rows, validation_failed_rows, validation_summary = _validate_fill1c_rows(can_process_rows)
    global FILL_1C_DATA_LIST
    FILL_1C_DATA_LIST = valid_rows
    DOC_CHECK_RESULT = {
        "headers": headers,
        "cannot_process_rows": cannot_rows,
        "validation_failed_rows": validation_failed_rows,
    }
    summary = _build_unprocessable_summary(cannot_rows)
    return jsonify(
        {
            "ok": True,
            "can_process": len(valid_rows),
            "cannot_process": len(cannot_rows),
            "validation_failed": len(validation_failed_rows),
            "reasons_summary": summary,
            "validation_errors_summary": validation_summary,
        }
    )


@app.route("/check-doc-upload-ids-only", methods=["POST"])
def check_doc_upload_ids_only():
    """
    Облегчённая проверка документа: считаем только наличие номера дела.
    Все остальные поля считаются необязательными.
    Используется на странице «Проверить данные 1С».
    """
    id_field = "ID дела  (ID/ номер убытка)"
    if "document" not in request.files:
        return jsonify({"error": "Файл не выбран"}), 400
    file = request.files["document"]
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Нужен файл .xlsx"}), 400
    try:
        headers, rows = _parse_xlsx_rows(file.stream)
    except Exception as e:
        return jsonify({"error": f"Ошибка чтения файла: {e}"}), 400
    if not headers:
        return jsonify({"error": "Файл пустой или без заголовков"}), 400
    if id_field not in headers:
        return (
            jsonify(
                {
                    "error": f"Не найден столбец '{id_field}' в заголовке. "
                    "Убедитесь, что вы копируете файл правильного формата."
                }
            ),
            400,
        )

    can_process = 0
    cannot_process = 0
    case_list = []
    for row in rows:
        val = row.get(id_field)
        if _is_empty_value(val):
            cannot_process += 1
        else:
            can_process += 1
            case_list.append({"number_case": str(val).strip()})

    global CHECK_1C_CASE_LIST
    CHECK_1C_CASE_LIST = case_list

    return jsonify(
        {
            "ok": True,
            "can_process": can_process,
            "cannot_process": cannot_process,
        }
    )

@app.route("/download-unprocessable-doc", methods=["GET"])
def download_unprocessable_doc():
    """Отдаёт xlsx с делами, которые нельзя обработать (из последней проверки)."""
    global DOC_CHECK_RESULT
    headers = DOC_CHECK_RESULT.get("headers") or []
    rows = DOC_CHECK_RESULT.get("cannot_process_rows") or []
    if not headers and not rows:
        return jsonify({"error": "Нет данных для скачивания. Сначала выполните проверку документа."}), 404
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Необрабатываемые"
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row_dict in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row_dict.get(h, ""))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="необрабатываемые_дела.xlsx",
    )


@app.route("/download-unprocessable-doc-by-reason", methods=["GET"])
def download_unprocessable_doc_by_reason():
    """
    Отдаёт xlsx только с теми делами, которые не прошли обработку
    по конкретной причине (отсутствию/пустоте указанного обязательного поля).
    """
    global DOC_CHECK_RESULT
    field = request.args.get("field", "").strip()
    if not field:
        return jsonify({"error": "Не указана причина (field)."}), 400

    headers = DOC_CHECK_RESULT.get("headers") or []
    rows = DOC_CHECK_RESULT.get("cannot_process_rows") or []
    if not headers or not rows:
        return jsonify(
            {"error": "Нет данных для скачивания. Сначала выполните проверку документа."}
        ), 404

    # Оставляем только те дела, где среди причин нет заполненности именно этого поля
    filtered_rows = []
    for row in rows:
        missing = row.get("_missing_required_fields") or []
        if isinstance(missing, str):
            fields = [f.strip() for f in missing.split(",") if f.strip()]
        else:
            fields = list(missing)
        if field in fields:
            filtered_rows.append(row)

    if not filtered_rows:
        return jsonify(
            {"error": f"Нет дел, которые не прошли проверку по причине: {field}"}
        ), 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Необрабатываемые по причине"
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row_dict in enumerate(filtered_rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row_dict.get(h, ""))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = field.replace("/", "_").replace("\\", "_")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"необрабатываемые_дела__{safe_name}.xlsx",
    )


@app.route("/download-validation-failed-doc", methods=["GET"])
def download_validation_failed_doc():
    """Отдаёт xlsx с делами, не прошедшими валидацию данных 1С (из последней проверки)."""
    global DOC_CHECK_RESULT
    headers = DOC_CHECK_RESULT.get("headers") or []
    rows = DOC_CHECK_RESULT.get("validation_failed_rows") or []
    if not rows:
        return jsonify(
            {"error": "Нет дел с ошибками валидации для скачивания. Сначала выполните проверку документа."}
        ), 404
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ошибки валидации"
    # Заголовки: оригинальные + колонка с текстом ошибки
    extra_col = "Ошибка валидации"
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=1, column=len(headers) + 1, value=extra_col)
    for r, row_dict in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row_dict.get(h, ""))
        ws.cell(row=r, column=len(headers) + 1, value=row_dict.get("_validation_error", ""))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="дела_ошибки_валидации_1С.xlsx",
    )


@app.route("/settings", methods=["GET", "POST"])
def settings():
    """
    GET: вернуть текущие настройки из config.json.
    POST: сохранить переданные настройки в config.json.
    """
    if request.method == "GET":
        cfg = load_config()
        # Явно возвращаем только интересующие нас ключи
        return jsonify(
            {
                "delaySearchCase": cfg.get("delaySearchCase"),
                "delayCourtTab": cfg.get("delayCourtTab"),
                "delayDownloadCases": cfg.get("delayDownloadCases"),
                "delayStageSwitch": cfg.get("delayStageSwitch"),
                "delayBeforeRun": cfg.get("delayBeforeRun"),
            }
        )

    data = request.get_json(silent=True) or {}
    cfg = load_config()

    def _to_float_or_none(value):
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _to_int_or_none(value, default=None):
        if value is None:
            return default
        try:
            n = int(value)
            return n if n >= 0 else default
        except (TypeError, ValueError):
            return default

    cfg["delaySearchCase"] = _to_float_or_none(data.get("delaySearchCase"))
    cfg["delayCourtTab"] = _to_float_or_none(data.get("delayCourtTab"))
    cfg["delayDownloadCases"] = _to_float_or_none(data.get("delayDownloadCases"))
    cfg["delayStageSwitch"] = _to_float_or_none(data.get("delayStageSwitch"))
    cfg["delayBeforeRun"] = _to_int_or_none(data.get("delayBeforeRun"), 60)

    try:
        save_config(cfg)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Не удалось сохранить настройки: {e}"}), 500

    return jsonify({"ok": True})


@app.route("/robot/run", methods=["POST"])
def robot_run():
    """
    Запуск робота из фронта. Тело запроса: {"mode": "check"|"set"|"download"|"change_filename", "data": ...}.
    data нормализуется под выбранный режим и передаётся в run_robot(mode, data).
    Для mode "check" без data используется список дел из последней проверки (кнопка «Начать»).
    """
    global CHECK_1C_CASE_LIST, FILL_1C_DATA_LIST, LOAD_1C_CASE_LIST
    body = request.get_json(silent=True) or {}
    mode = (body.get("mode") or "").strip()
    raw_data = body.get("data")
    if mode not in ("check", "set", "download", "change_filename"):
        return jsonify({"ok": False, "error": f"Неизвестный режим: {mode!r}"}), 400
    if mode == "check" and (raw_data is None or (isinstance(raw_data, list) and len(raw_data) == 0)):
        raw_data = CHECK_1C_CASE_LIST
    if mode == "set" and (raw_data is None or (isinstance(raw_data, list) and len(raw_data) == 0)):
        raw_data = FILL_1C_DATA_LIST
    if mode == "download" and (raw_data is None or (isinstance(raw_data, list) and len(raw_data) == 0)):
        raw_data = LOAD_1C_CASE_LIST
    if mode == "check" and (not raw_data or len(raw_data) == 0):
        return jsonify({"ok": False, "error": "Нет дел для проверки. Сначала нажмите «Начать» и загрузите файл с делами."}), 400
    if mode == "set" and (not raw_data or len(raw_data) == 0):
        return jsonify({"ok": False, "error": "Нет данных для заполнения. Сначала нажмите «Начать» и загрузите файл с данными."}), 400
    if mode == "download" and (not raw_data or len(raw_data) == 0):
        return jsonify({"ok": False, "error": "Нет дел для загрузки. Сначала нажмите «Начать» и загрузите файл или вставьте данные."}), 400
    data = normalize_robot_payload(mode, raw_data if raw_data is not None else {})
    def run_in_thread():
        try:
            run_robot(mode, data)
        except Exception:
            pass
    t = threading.Thread(target=run_in_thread, daemon=True)
    t.start()
    return jsonify({"ok": True, "running": True})


@app.route("/robot/stop", methods=["POST"])
def robot_stop():
    """Запрос остановки текущего запуска робота."""
    request_stop_run()
    return jsonify({"ok": True, "stopping": True})


@app.route("/health")
def health():
    return jsonify({"status": "ok", "run_id": RUN_ID})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, debug=True)

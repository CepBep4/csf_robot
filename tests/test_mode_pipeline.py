import sys
import types
import unittest
from io import BytesIO
from unittest.mock import patch

import openpyxl

import robot
import robot_last_results
import server


class TestModePipeline(unittest.TestCase):
    def test_interface_start_runs_prod_mode_and_returns_step_logs(self):
        mode_key = "09_Ожидание решения суда"
        fake_mode_all = types.ModuleType("worker.modeAll")
        fake_mode_all.MODE_PROD_MAP = {
            mode_key: ["Открыть дело", "Заполнить вкладку суд"],
        }

        def _fake_run_prod_mode(case, mode_key, cooldown=0):
            return [
                {"step": "Открыть дело", "ok": True, "message": "Дело открыто"},
                {"step": "Заполнить вкладку суд", "ok": True, "message": "Суд заполнен"},
            ]

        fake_mode_all.run_prod_mode = _fake_run_prod_mode

        payload = [{
            "number_case": "18-031628",
            "setInfo": mode_key,
            "validateBeforeRun": False,
        }]

        with patch.dict(sys.modules, {"worker.modeAll": fake_mode_all}):
            results = robot.interface_start("set", payload)

        self.assertEqual(len(results), 1)
        self.assertTrue(results[0]["ok"])
        self.assertEqual(results[0]["mode"], mode_key)
        self.assertIn("steps", results[0])
        self.assertEqual(len(results[0]["steps"]), 2)
        self.assertEqual(results[0]["steps"][0]["step"], "Открыть дело")

    def test_download_last_run_results_contains_steps_logs_column(self):
        robot_last_results.set_last_run(
            "set",
            [{
                "number_case": "18-031628",
                "ok": True,
                "message": "Успешно",
                "steps": [
                    {"step": "Открыть дело", "ok": True, "message": "Дело открыто"},
                    {"step": "Заполнить вкладку суд", "ok": True, "message": "Суд заполнен"},
                ],
            }],
        )

        client = server.app.test_client()
        resp = client.get("/robot/download-last-run-results")
        self.assertEqual(resp.status_code, 200)

        wb = openpyxl.load_workbook(BytesIO(resp.data))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=4).value, "Логи этапов")
        logs_cell = ws.cell(row=2, column=4).value
        self.assertIsInstance(logs_cell, str)
        self.assertIn("Открыть дело", logs_cell)


if __name__ == "__main__":
    unittest.main()
